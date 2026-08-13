"""ייבוא מוזמנים מקובץ Excel/CSV: קריאה, זיהוי עמודות חכם וולידציה.

זיהוי העמודות מבוסס על מילות-מפתח בכותרות (עברית/אנגלית), כדי לתמוך
בקבצים שהמשתמשים מביאים בפורמטים שונים.

בנוסף, `parse_freeform_text` מפענח רשימת טקסט חופשי (הדבקה מ-WhatsApp/אקסל/
כל מקום) שורה-שורה: מזהה טלפון, כמות אנשים, רמז "משפחת…", ומשאיר את השם.
הפענוח דטרמיניסטי לגמרי (regex/היוריסטיקה) — אין קריאות LLM.
"""
import csv
import io
import re
from typing import Optional

from openpyxl import load_workbook

from app.validators import normalize_israeli_phone

# מילות מפתח לזיהוי כל עמודה (בכותרת). הבדיקה: האם הכותרת מכילה אחת מהן.
COLUMN_KEYWORDS = {
    "full_name": ["שם מלא", "שם", "name", "מוזמן"],
    "phone": ["טלפון", "נייד", "פלאפון", "פלאפו", "phone", "mobile", "טל"],
    "side": ["צד", "side"],
    "group_type": ["קבוצה", "קבוצת", "group", "שיוך"],
    "party_size": ["כמות", "אנשים", "מוזמנים", "size", "count"],
    # seating_notes חייב לבוא **לפני** notes_raw: detect_columns משייך כל
    # עמודה לשדה הראשון שתופס אותה, ולכן עמודה "הערות הושבה" צריכה להיתפס
    # כאן ולא ע"י המילה הכללית "הערות" שמתחתיה.
    "seating_notes": ["הערות הושבה", "הערת הושבה", "הושבה", "seating"],
    "notes_raw": ["הערה", "הערות", "notes", "מגבל"],
}

SIDE_VALUE_MAP = {
    # ערכי "צד" לחתונה/חינה (חתן/כלה) ולסוגי אירוע אחרים (אב/אם, א׳/ב׳) —
    # פנימית כולם נשמרים כ-groom/bride/shared, והתווית המוצגת נשאבת מ-
    # eventTypes.ts לפי event_type (ראה sideLabel/sidePhrase).
    "חתן": "groom",
    "groom": "groom",
    "אב": "groom",
    "א׳": "groom",
    "כלה": "bride",
    "bride": "bride",
    "אמא": "bride",
    "אם": "bride",
    "ב׳": "bride",
    "משותף": "shared",
    "שני": "shared",
    "shared": "shared",
}

GROUP_VALUE_MAP = {
    # חתונה/חינה — קטגוריות ספציפיות קודם, כדי שלא "יבלעו" ע"י ההתאמה הכללית
    # של "משפחה"/"אב"/"אם" למטה (הראשון שמתאים בסדר האיטרציה מנצח).
    "משפחה קרובה": "close_family",
    "קרובה": "close_family",
    "משפחה רחוקה": "extended_family",
    "רחוקה": "extended_family",
    "חברים": "friends",
    "חבר": "friends",
    "עבודה": "work",
    "work": "work",
    "צבא": "army",
    "לימודים": "studies",
    "ילדות": "childhood",
    "שכנים": "neighbors",
    # בר/בת מצווה
    "משפחת האב": "family_father",
    "משפחת אב": "family_father",
    "משפחת האם": "family_mother",
    "משפחת אם": "family_mother",
    "כיתה": "class",
    "חוגים": "staff_clubs",
    "צוות": "staff_clubs",
    # אב/אם עצמאיים (אחרי הצירופים הארוכים למעלה) — "אב"/"אבא" ו"אם"/"אמא"
    "אבא": "family_father",
    "אב": "family_father",
    "אמא": "family_mother",
    "אם": "family_mother",
    # ברית/משפחתי/חינה — כללי, אחרי כל הצירופים הספציפיים למעלה
    "משפחה": "family",
    # אירוע עסקי
    "עובדים": "employees",
    "עובד": "employees",
    "לקוחות": "clients",
    "לקוח": "clients",
    "ספקים": "suppliers",
    "ספק": "suppliers",
    "הנהלה": "management",
    "מנהל": "management",
    "שותפים": "partners",
    "שותף": "partners",
}


_FRIENDLY_READ_ERROR = "לא הצלחנו לקרוא את הקובץ. בדקו שהקובץ תקין ונסו שוב."

# סדר ניסיונות קידוד ל-CSV: UTF-8 (עם/בלי BOM) קודם כי זו ברירת המחדל של
# Google Sheets ורוב הכלים המודרניים; cp1255/Windows-1255 כגיבוי כי זה
# הקידוד שאקסל בעברית על Windows עדיין מייצא אליו לפעמים ("CSV (Comma
# delimited)") — בלי הגיבוי הזה שמות בעברית הופכים לג'יבריש בשקט, בלי שגיאה.
_CSV_ENCODINGS = ["utf-8-sig", "cp1255", "iso-8859-8"]


def _decode_csv_text(content: bytes) -> str:
    """מנסה כמה קידודים בסדר עדיפות; רק אם כולם נכשלים, נופל ל-UTF-8 עם
    replace (עדיף טקסט חלקי קריא על פני קריסה, אבל זה המוצא האחרון)."""
    for enc in _CSV_ENCODINGS:
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[list]]:
    """מחזיר (כותרות, שורות) מקובץ CSV או XLSX.

    כל שגיאת קריאה (קובץ פגום, לא-זיפ, XLSX שקרס, קידוד לא צפוי) נתפסת פה
    והופכת ל-ValueError עם הודעה בעברית שהמשתמש יכול להבין ולפעול לפיה —
    לא מותר שקובץ פגום יגרום ל-500 גולמי.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            text = _decode_csv_text(content)
            reader = list(csv.reader(io.StringIO(text)))
        except Exception:
            raise ValueError(_FRIENDLY_READ_ERROR)
        if not reader:
            return [], []
        headers = [str(h).strip() for h in reader[0]]
        rows = [list(r) for r in reader[1:]]
        return headers, rows
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            raise ValueError(_FRIENDLY_READ_ERROR)
        if not all_rows:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
        rows = [list(r) for r in all_rows[1:]]
        return headers, rows
    raise ValueError("פורמט קובץ לא נתמך. נא להעלות קובץ .xlsx או .csv")


def detect_columns(headers: list[str]) -> dict:
    """ממפה שדה -> אינדקס עמודה, לפי מילות מפתח בכותרת.

    עובד לפי סדר עדיפויות (השדות ב-COLUMN_KEYWORDS), ומוודא שכל עמודה
    משויכת לכל היותר לשדה אחד — כדי למנוע התנגשויות (למשל "מספר טלפון"
    שמכיל גם 'טלפון' וגם מספר).
    """
    mapping: dict[str, Optional[int]] = {field: None for field in COLUMN_KEYWORDS}
    used: set[int] = set()
    for field, keywords in COLUMN_KEYWORDS.items():
        for idx, header in enumerate(headers):
            if idx in used:
                continue
            h = (header or "").strip().lower()
            if not h:
                continue
            if any(kw.lower() in h for kw in keywords):
                mapping[field] = idx
                used.add(idx)
                break
    return mapping


def _cell(row: list, idx: Optional[int]) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _map_side(raw: str) -> str:
    key = raw.strip().lower()
    for text, value in SIDE_VALUE_MAP.items():
        if text.lower() in key:
            return value
    return "shared"


def _map_group(raw: str) -> str:
    key = raw.strip().lower()
    for text, value in GROUP_VALUE_MAP.items():
        if text.lower() in key:
            return value
    return "other"


def build_preview(headers: list[str], rows: list[list], mapping: dict) -> dict:
    """מייצר תצוגה מקדימה עם ולידציה לכל שורה."""
    preview_rows = []
    valid_count = 0

    for i, row in enumerate(rows):
        # מדלגים על שורות ריקות לגמרי
        if not any(str(c).strip() for c in row if c is not None):
            continue

        full_name = _cell(row, mapping.get("full_name"))
        phone_raw = _cell(row, mapping.get("phone"))
        side = _map_side(_cell(row, mapping.get("side")))
        group_type = _map_group(_cell(row, mapping.get("group_type")))
        notes_raw = _cell(row, mapping.get("notes_raw")) or None
        seating_notes = _cell(row, mapping.get("seating_notes")) or None

        party_raw = _cell(row, mapping.get("party_size"))
        try:
            party_size = int(float(party_raw)) if party_raw else 1
            if party_size < 1:
                party_size = 1
        except ValueError:
            party_size = 1

        errors = []
        if not full_name:
            errors.append("חסר שם")

        phone = phone_raw
        try:
            phone = normalize_israeli_phone(phone_raw)
        except ValueError:
            errors.append("טלפון לא תקין")

        is_valid = len(errors) == 0
        if is_valid:
            valid_count += 1

        preview_rows.append(
            {
                "row_number": i + 2,  # +2: שורה 1 = כותרות, אינדקס מ-0
                "full_name": full_name,
                "phone": phone,
                "side": side,
                "group_type": group_type,
                "party_size": party_size,
                "notes_raw": notes_raw,
                "seating_notes": seating_notes,
                "valid": is_valid,
                "errors": errors,
            }
        )

    return {
        "detected_columns": {
            field: (headers[idx] if idx is not None and idx < len(headers) else None)
            for field, idx in mapping.items()
        },
        "rows": preview_rows,
        "total": len(preview_rows),
        "valid_count": valid_count,
        "invalid_count": len(preview_rows) - valid_count,
    }


# ---------------------------------------------------------------------------
# פענוח טקסט חופשי (הדבקת רשימה מ-WhatsApp / אקסל / כל מקור)
# ---------------------------------------------------------------------------

# רצף שנראה כמו טלפון ישראלי: מתחיל ב-0 (מקומי) או +972/972 (בינ"ל), עם עד
# ספרה אחת בין כל שתי ספרות (רווח/מקף מותרים כמפרידים). מספר תקין הוא 9-10
# ספרות מקומי (אחרי הסרת 972 והחלפתו ב-0) — ולכן חוזרים על "ספרה נוספת" 8-9
# פעמים בדיוק, לא ללא הגבלה. ה-bound הזה קריטי: בלעדיו רצף כמו "...1234567 5"
# (טלפון ואז כמות בשורה) היה "בולע" את ה-5 כאילו הוא הספרה ה-11 של הטלפון.
_PHONE_RE = re.compile(r"(?:\+?972(?:[\s\-]?\d){8,9}|0(?:[\s\-]?\d){8,9})")

# כמות מפורשת: "5 אנשים", "5 מוזמנים", "5 נפשות", "5 איש/אורחים"
_COUNT_WORD_RE = re.compile(r"(\d+)\s*(?:אנשים|מוזמנים|נפשות|איש|אורחים)")
# כמות בסוגריים: "(5)" או "[5]"
_COUNT_PAREN_RE = re.compile(r"[\(\[]\s*(\d+)\s*[\)\]]")
# כמות עם x/×/*: "x5", "X 5", "*5"
_COUNT_X_RE = re.compile(r"(?:^|\s)[xX*×]\s*(\d+)\b")
# מספר בודד בקצה השורה (אחרי שהוסר הטלפון): "דנה 2"
_COUNT_TRAIL_RE = re.compile(r"(?:^|\s)(\d+)\s*$")

# רמז "משפחת …" / "משפחה של …" → קבוצת משפחה קרובה. זה רק תיוג group_type —
# **לא** משפיע על הכמות (ראה _parse_quantity): "משפחת לוי" בלי כמות מפורשת
# נשארת עם כמות לא-ידועה, בדיוק כמו כל שורה אחרת בלי כמות.
_FAMILY_RE = re.compile(r"^\s*משפח[הת]\b")

# ---------------------------------------------------------------------------
# זיהוי כמות בעברית — מילים ומספרים
# ---------------------------------------------------------------------------
# ממיר מילת-מספר עברית (זכר/נקבה) לערך שלה. משמש גם למספר "כמות ישירה"
# (למשל "שלושה" לבד = 3 אנשים) וגם כמספר ילדים ("שני ילדים" = 2 ילדים).
_NUMBER_WORDS: dict[str, int] = {
    "אחד": 1, "אחת": 1,
    "שניים": 2, "שתיים": 2, "שני": 2, "שתי": 2,
    "שלושה": 3, "שלוש": 3,
    "ארבעה": 4, "ארבע": 4,
    "חמישה": 5, "חמש": 5,
    "שישה": 6, "שש": 6,
    "שבעה": 7, "שבע": 7,
    "שמונה": 8,
    "תשעה": 9, "תשע": 9,
    "עשרה": 10, "עשר": 10,
}
# ממוינות מהארוך לקצר כדי שההתאמה לא תיעצר על תת-מחרוזת ("שני" בתוך "שניים").
_NUM_WORD_ALT = "|".join(sorted(_NUMBER_WORDS, key=len, reverse=True))
_NUM_TOKEN_RE = rf"(?:\d+|{_NUM_WORD_ALT})"

_CHILD_WORD_RE = r"ילד(?:ים|ה|ות)?"
_COUPLE_WORD_RE = r"(?:זוג|זוגי|שני\s+מבוגרים)"

# "זוג + ילד" / "זוג עם שני ילדים" — שני מבוגרים ועוד ילדים, מחוברים ב-+/עם/ו.
_COMBINED_QTY_RE = re.compile(
    rf"{_COUPLE_WORD_RE}\s*(?:\+|עם|ו-?)\s*"
    rf"(?:(?P<before>{_NUM_TOKEN_RE})\s+)?{_CHILD_WORD_RE}(?:\s+(?P<after>{_NUM_TOKEN_RE}))?"
)
# ילד/ה בודד/ים, עם מספר לפני או אחרי או בלי (ילד בודד = ילד אחד).
_CHILD_QTY_RE = re.compile(
    rf"(?:(?P<before>{_NUM_TOKEN_RE})\s+)?{_CHILD_WORD_RE}(?:\s+(?P<after>{_NUM_TOKEN_RE}))?"
)
_COUPLE_ALONE_RE = re.compile(rf"\b{_COUPLE_WORD_RE}\b")
_SINGLE_ALONE_RE = re.compile(r"\b(?:יחיד|בודד|אח[דת](?:\s+מגיע\w*)?)\b")
# מילת-מספר עברית עצמאית (לא צמודה ל"ילד") = כמות אנשים ישירה: "שלושה" -> 3.
_BARE_NUM_WORD_RE = re.compile(rf"(?:^|\s)(?P<w>{_NUM_WORD_ALT})(?:\s|$)")


def _num_token_value(token: Optional[str]) -> Optional[int]:
    if not token:
        return None
    token = token.strip()
    if token.isdigit():
        return int(token)
    return _NUMBER_WORDS.get(token)


def _parse_quantity(working: str) -> tuple[Optional[str], Optional[int], str]:
    """מזהה ביטוי כמות בשורה ומסירה אותו ממנה.

    מחזיר (guest_count_text, party_size, working_without_quantity):
    guest_count_text = בדיוק מה שהמשתמש כתב (או None אם לא זוהתה כמות),
    party_size = מספר האנשים בפועל שחושב מהביטוי (או None — ולא 1! — אם
    לא זוהתה כמות; חל איסור לנחש "יחיד" כברירת מחדל, ראה `parse_freeform_text`).

    סדר הניסיונות הוא לפי ספציפיות: קודם צירוף "זוג+ילדים", אחר כך "ילדים"
    לבד (מוזמן ראשי + הילדים), רק אז כמויות ישירות (מספר/זוג/יחיד/מילת-מספר).
    """
    m = _COMBINED_QTY_RE.search(working)
    if m:
        children = _num_token_value(m.group("before")) or _num_token_value(m.group("after")) or 1
        text = m.group(0).strip()
        rest = working[: m.start()] + " " + working[m.end():]
        return text, 2 + children, rest

    m = _CHILD_QTY_RE.search(working)
    if m:
        children = _num_token_value(m.group("before")) or _num_token_value(m.group("after")) or 1
        text = m.group(0).strip()
        rest = working[: m.start()] + " " + working[m.end():]
        return text, 1 + children, rest

    for rx in (_COUNT_WORD_RE, _COUNT_PAREN_RE, _COUNT_X_RE):
        m = rx.search(working)
        if m:
            text = m.group(0).strip()
            rest = working[: m.start()] + " " + working[m.end():]
            return text, int(m.group(1)), rest

    m = _COUPLE_ALONE_RE.search(working)
    if m:
        rest = working[: m.start()] + " " + working[m.end():]
        return m.group(0).strip(), 2, rest

    m = _SINGLE_ALONE_RE.search(working)
    if m:
        rest = working[: m.start()] + " " + working[m.end():]
        return m.group(0).strip(), 1, rest

    m = _BARE_NUM_WORD_RE.search(working)
    if m:
        value = _num_token_value(m.group("w"))
        rest = working[: m.start()] + " " + working[m.end():]
        return m.group("w").strip(), value, rest

    m = _COUNT_TRAIL_RE.search(working)
    if m:
        rest = working[: m.start()] + " " + working[m.end():]
        return m.group(1).strip(), int(m.group(1)), rest

    return None, None, working


def _clean_name(text: str) -> str:
    """מנקה מהשם שאריות מפרידים (מקפים/פסיקים/נקודתיים) ורווחים כפולים."""
    text = re.sub(r"[\-–—,:;|]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" -–—,:;|\t")


def parse_line(line: str) -> dict:
    """מפענחת **שורה אחת בלבד** — יחידת הפענוח האטומית של הדבקת רשימה.

    זו הערבות המבנית לבידוד בין שורות: הפונקציה מקבלת ``str`` בודד ומחזירה
    רק מידע שנמצא בתוך אותה מחרוזת. אין לה גישה לשורות אחרות, אין לה
    state שמצטבר בין קריאות (כל משתנה מקומי לפונקציה), ואין לה שום דרך
    "לדעת" שקיימת שורה קודמת/הבאה — ולכן טלפון/כמות של שורה אחת פיזית
    לא יכולים "לדלוף" לשורה אחרת. `parse_freeform_text` קורא לה בלולאה
    ומוסיף מעליה רק דברים שמטבעם חוצי-שורות (מספור, זיהוי כפילות מול
    שורות אחרות/מוזמנים קיימים) — לא מידע על השורה עצמה.

    מחזירה: full_name, phone, phone_warn, guest_count_text, party_size
    (None אם לא זוהתה כמות בשורה — לעולם לא מוחלף בברירת מחדל כאן), group_type.
    """
    working = line

    # 1) טלפון — מזהים, מנרמלים, ומסירים מהשורה
    phone = ""
    phone_warn: Optional[str] = None
    m = _PHONE_RE.search(working)
    if m:
        candidate = m.group(0)
        working = working[: m.start()] + " " + working[m.end():]
        try:
            phone = normalize_israeli_phone(candidate)
        except ValueError:
            phone_warn = "טלפון לא תקין"

    # 2) כמות — מזהים ומסירים מהשורה (ראו _parse_quantity). קורא ל-working
    # הנוכחי בלבד — לא לשום דבר שהצטבר משורה קודמת.
    guest_count_text, party_size, working = _parse_quantity(working)

    # 3) רמז משפחה — תיוג group_type בלבד, לא כמות ולא קשר לשורות אחרות
    is_family = bool(_FAMILY_RE.search(line))
    group_type = "close_family" if is_family else "other"

    # 4) שם — מה שנשאר אחרי הסרת טלפון+כמות
    full_name = _clean_name(working)

    return {
        "full_name": full_name,
        "phone": phone,
        "phone_warn": phone_warn,
        "guest_count_text": guest_count_text,
        "party_size": party_size,
        "group_type": group_type,
    }


def parse_freeform_text(
    text: str,
    existing_keys: Optional[set] = None,
    assume_single_if_no_count: bool = False,
) -> dict:
    """מפענח רשימת טקסט חופשי לשורות מוזמנים — מבנה זהה ל-`build_preview`.

    כל שורה עוברת דרך `parse_line` בבידוד מוחלט (ראו התיעוד שם). הפונקציה
    הזו מוסיפה מעליה **רק** את מה שמטבעו חוצה-שורות: מספור (`row_number`),
    וזיהוי כפילות (`duplicate`) מול שורות אחרות באותה הדבקה ומול מוזמני
    האירוע (`existing_keys`) — שני הדברים היחידים שבאמת אמורים להשוות בין
    שורות. שום מידע אחר (שם/טלפון/כמות) לא זולג בין שורות.

    כלל ברזל: **לא מנחשים כמות שלא נכתבה.** אם `parse_line` לא זיהה ביטוי
    כמות, `guest_count_text`/`party_size` נשארים None ומתווספת אזהרה חוסמת
    ("חסרה כמות") — לא ברירת מחדל שקטה ל"יחיד". היוצא מהכלל היחיד: כש-
    `assume_single_if_no_count=True` (משמש לזרימת ייבוא אנשי קשר, ששם כל
    שורה היא כבר איש קשר בודד בוודאות — 1 הוא עובדה נתונה, לא ניחוש).

    ולידציה: `errors` חוסמים (חסר שם) לעומת `warnings` חוסמי-ברירת-מחדל
    (חסרה כמות) או לא-חוסמים (חסר טלפון / טלפון לא תקין / כפילות).
    """
    existing_keys = existing_keys or set()
    preview_rows = []
    valid_count = 0
    seen_keys: set[str] = set()
    row_no = 0

    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        row_no += 1

        parsed = parse_line(line)
        full_name = parsed["full_name"]
        phone = parsed["phone"]
        phone_warn = parsed["phone_warn"]
        guest_count_text = parsed["guest_count_text"]
        party_size = parsed["party_size"]
        group_type = parsed["group_type"]

        count_missing = party_size is None
        if count_missing and assume_single_if_no_count:
            party_size = 1

        # ולידציה
        errors: list[str] = []
        warnings: list[str] = []
        if not full_name:
            errors.append("חסר שם")
        if count_missing and not assume_single_if_no_count:
            warnings.append("חסרה כמות")
        if phone_warn:
            warnings.append(phone_warn)
        elif not phone:
            warnings.append("חסר טלפון")

        # כפילות — מפתח לפי טלפון (מדויק) או שם (fallback). זה היחיד שמותר
        # לו להשוות בין שורות (השוואה, לא העברת מידע).
        key = phone or (full_name.lower() if full_name else "")
        duplicate = bool(key and (key in seen_keys or key in existing_keys))
        if duplicate:
            warnings.append("כפילות")
        if key:
            seen_keys.add(key)

        is_valid = len(errors) == 0 and party_size is not None
        if is_valid:
            valid_count += 1

        preview_rows.append(
            {
                "row_number": row_no,
                "full_name": full_name,
                "phone": phone,
                "side": "shared",
                "group_type": group_type,
                "guest_count_text": guest_count_text,
                # 0 = כמות טרם זוהתה (לא כמות אפס בפועל — כמות אמיתית תמיד ≥1).
                # sentinel ולא None כדי לשמור על party_size:int יציב בחוזה ה-API.
                "party_size": party_size if party_size is not None else 0,
                "notes_raw": None,
                "seating_notes": None,
                "valid": is_valid,
                "errors": errors,
                "warnings": warnings,
                "duplicate": duplicate,
            }
        )

    return {
        "detected_columns": {},
        "rows": preview_rows,
        "total": len(preview_rows),
        "valid_count": valid_count,
        "invalid_count": len(preview_rows) - valid_count,
    }
